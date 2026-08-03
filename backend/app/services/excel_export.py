"""Excel 导出服务：基于 openpyxl 生成测试用例工作簿。"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# 导出列定义（与 Excel 模板一致）
EXPORT_HEADERS = [
    "编号",
    "优先级",
    "模块",
    "功能",
    "测试点",
    "前置条件",
    "测试步骤",
    "测试数据",
    "预期结果",
    "备注",
]

HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
THIN_SIDE = Side(style="thin", color="B4C6E7")
CELL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def build_testcase_excel(data_rows: list[list[object]]) -> bytes:
    """生成测试用例 Excel 文件，返回字节内容。

    data_rows 为二维列表，顺序与 EXPORT_HEADERS 一致。
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "测试用例"

    # 表头
    sheet.append(EXPORT_HEADERS)
    for cell in sheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = CELL_BORDER
    sheet.row_dimensions[1].height = 24

    # 数据行
    for row in data_rows:
        sheet.append(row)

    # 数据单元格样式：边框 + 自动换行
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.border = CELL_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    # 冻结首行
    sheet.freeze_panes = "A2"

    # 自动设置列宽：按内容最大显示宽度（中文按 2 个字符计），限制在 12-50
    for index, header in enumerate(EXPORT_HEADERS, start=1):
        max_width = _display_width(header)
        for cell in sheet.iter_rows(
            min_row=2, min_col=index, max_col=index, values_only=True
        ):
            value = cell[0]
            if value:
                max_width = max(max_width, _display_width(str(value)))
        sheet.column_dimensions[get_column_letter(index)].width = min(
            max(max_width + 4, 12), 50
        )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _display_width(text: str) -> int:
    """估算文本显示宽度：中文等宽字符按 2 计算。"""
    return sum(2 if ord(char) > 127 else 1 for char in text)
